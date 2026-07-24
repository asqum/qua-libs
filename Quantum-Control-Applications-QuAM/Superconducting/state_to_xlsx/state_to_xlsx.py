"""
使用方法:
1. 將 state.json 放入 Design 中該 QPU 的資料夾，如果沒有則自己創建一個
2. 將電阻預測 qubit 頻率的 csv 檔（位於 Chip_candidate_resistances），
   改名為 R_to_fq.csv 後複製至此資料夾
3. 依序跑下面 cell（cell 4 會產生並儲存兩張散布圖）
4. 如果是要更新同一次量測結果的話，將 new_meas 改為 False
"""
# %%
import json
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd

file_path = r"\\10.21.19.203\QCTSS-Shared\asqcmeasurement\Sample property\new_sorting\QPU\10FQ9FCv2\10FQ9FCv2#12_251222"
new_meas = True

TIME_S_KEYS = {"T1", "T2", "T1_dev", "T2_dev", "T2ramsey", "T2ramsey_dev"}
SKIP_KEYS = {"EPC", "EPC_incoh", "EPG_incoh", "iso_EPC", "iso_EPG"}
KEY_LABELS = {"EPG": "RB fidelity"}
MEAS_PATTERN = re.compile(r"meas#(\d+)\.xlsx", re.IGNORECASE)
XLSX_ROW_MAP = {
    "T1": "T1 #100 (us)",
    "T1_dev": "T1_error (us)",
    "T2": "T2_echo #100 (us)",
    "T2_dev": "T2_echo_error (us)",
    "T2ramsey": "T2_Ramsey #100 (us)",
    "T2ramsey_dev": "T2_Ramsey_error (us)",
    "Teff_mK": "Temperature #100 (mk)",
    "Teff_mK_dev": "Temperature_error #100 (mk)",
    "sweetspot_freq": "Qubit frequency (GHz)",
    "FluxPeriodV": "flux_period (V)",
    "bare_resonator_freq": "Bare resonant frequency (GHz)",
    "dressed_resonator_freq": "Dressed resonant frequency (|0>)(GHz)",
    "readout_fidelity": "Readout fidelity",
    "RB fidelity": "RB fidelity",
}


def qubit_sort_key(name: str) -> int:
    return int(re.search(r"\d+", name).group())


def format_value(key: str, value) -> str:
    if value is None or value == "":
        return ""
    if not isinstance(value, (int, float)):
        return ""
    if key in TIME_S_KEYS:
        return f"{value * 1e6:.2f}"
    if "freq" in key.lower():
        return f"{value / 1e9:.3f}"
    if key == "EPG":
        return f"{(1 - value) * 100:.2f}"
    if key == "readout_fidelity":
        return f"{value:.2f}"
    return f"{value:.2f}"


def load_qubit_extras(state_path: Path) -> pd.DataFrame:
    with state_path.open(encoding="utf-8") as f:
        state = json.load(f)

    qubits = sorted(state["qubits"], key=qubit_sort_key)
    extras_by_qubit = {
        q.upper(): state["qubits"][q].get("extras", {}) for q in qubits
    }
    all_keys = sorted(
        {k for extras in extras_by_qubit.values() for k in extras} - SKIP_KEYS
    )

    data = {
        KEY_LABELS.get(key, key): [
            format_value(key, extras_by_qubit[q].get(key)) for q in extras_by_qubit
        ]
        for key in all_keys
    }
    return pd.DataFrame(data, index=list(extras_by_qubit))


def meas_numbers(directory: Path) -> list[int]:
    numbers = []
    for path in directory.glob("meas*.xlsx"):
        match = MEAS_PATTERN.fullmatch(path.name)
        if match:
            numbers.append(int(match.group(1)))
    return numbers


def max_meas_number(directory: Path) -> int:
    numbers = meas_numbers(directory)
    return max(numbers) if numbers else 0


state_file = Path(file_path) / "state.json"
qubit_table = load_qubit_extras(state_file).T
print(qubit_table.to_string())

# %%
meas_dir = Path(file_path)
if new_meas:
    next_meas_number = max_meas_number(meas_dir) + 1
    new_meas_file = meas_dir / f"meas#{next_meas_number}.xlsx"
    shutil.copy2(meas_dir.parent / "dummy_meas.xlsx", new_meas_file)
    print(f"Created {new_meas_file}")
else:
    print("new_meas=False, skip creating meas file")

# %%
meas_dir = Path(file_path)
target_meas_file = meas_dir / f"meas#{max_meas_number(meas_dir)}.xlsx"
if not target_meas_file.exists():
    raise FileNotFoundError(f"No meas file found in {meas_dir}")

workbook = openpyxl.load_workbook(target_meas_file)
worksheet = workbook["Data"]

row_labels = {
    str(cell.value).strip(): row
    for row in range(1, worksheet.max_row + 1)
    if (cell := worksheet.cell(row, 1)).value
}
qubit_columns = {
    str(cell.value).strip().upper(): col
    for col in range(2, worksheet.max_column + 1)
    if (cell := worksheet.cell(14, col)).value
}

for source_key, xlsx_label in XLSX_ROW_MAP.items():
    if source_key not in qubit_table.index:
        continue
    row = row_labels.get(xlsx_label)
    if row is None:
        continue
    for qubit, value in qubit_table.loc[source_key].items():
        col = qubit_columns.get(qubit)
        if col is None or value in ("", None):
            continue
        worksheet.cell(row, col, float(value))

flux_row = row_labels.get("flux_period (V)")
tunable_row = row_labels.get("Tunable")
if flux_row and tunable_row and "FluxPeriodV" in qubit_table.index:
    for qubit, value in qubit_table.loc["FluxPeriodV"].items():
        col = qubit_columns.get(qubit)
        if col is None:
            continue
        worksheet.cell(tunable_row, col, value not in ("", None))

workbook.save(target_meas_file)
print(f"Filled {target_meas_file}")

# %%
import matplotlib.pyplot as plt

Q_PATTERN = re.compile(r"^Q(\d+)$", re.IGNORECASE)


def read_csv_q_values(csv_path: Path) -> dict[int, float]:
    values = {}
    for _, row in pd.read_csv(csv_path, header=None).iterrows():
        match = Q_PATTERN.fullmatch(str(row.iloc[0]).strip())
        if not match:
            continue
        try:
            values[int(match.group(1))] = float(row.iloc[1])
        except (TypeError, ValueError, IndexError):
            continue
    return values


def read_design_column(xlsx_path: Path, column_keyword: str) -> dict[int, float]:
    worksheet = openpyxl.load_workbook(xlsx_path, data_only=True).active
    target_col = None
    for row in range(1, 30):
        for col in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row, col).value
            if value and column_keyword in str(value):
                target_col = col
                start_row = row + 1
                break
        if target_col:
            break
    if not target_col:
        raise ValueError(f"Column containing '{column_keyword}' not found in {xlsx_path}")

    values = {}
    for row in range(start_row, worksheet.max_row + 1):
        match = Q_PATTERN.fullmatch(str(worksheet.cell(row, 1).value or "").strip())
        if not match:
            continue
        value = worksheet.cell(row, target_col).value
        if isinstance(value, (int, float)):
            values[int(match.group(1))] = float(value)
    return values


def read_meas_q_values(meas_path: Path, row_label: str, header_row: int) -> dict[int, float]:
    worksheet = openpyxl.load_workbook(meas_path, data_only=True)["Data"]
    data_row = next(
        row
        for row in range(1, worksheet.max_row + 1)
        if str(worksheet.cell(row, 1).value or "").strip() == row_label
    )
    values = {}
    for col in range(2, worksheet.max_column + 1):
        match = Q_PATTERN.fullmatch(str(worksheet.cell(header_row, col).value or "").strip())
        if not match:
            continue
        value = worksheet.cell(data_row, col).value
        if isinstance(value, (int, float)):
            values[int(match.group(1))] = float(value)
    return values


def plot_frequency_compare(
    q_numbers: list[int],
    series_a: dict[int, float],
    series_b: dict[int, float],
    label_a: str,
    label_b: str,
    title: str,
    output_path: Path,
    ylim: tuple[float, float],
) -> None:
    color_a, color_b = "#1f77b4", "#ff7f0e"
    x_vals = q_numbers
    y_a = [series_a.get(q) for q in q_numbers]
    y_b = [series_b.get(q) for q in q_numbers]

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.scatter(x_vals, y_a, s=100, c=color_a, label=label_a, zorder=3)
    ax.scatter(x_vals, y_b, s=100, c=color_b, label=label_b, zorder=3)
    ax.set_xlabel("Qubit number")
    ax.set_ylabel("Frequency (GHz)")
    ax.set_title(title)
    ax.set_ylim(*ylim)
    ax.set_xticks(q_numbers)
    ax.set_xticklabels([f"Q{q}" for q in q_numbers])
    ax.yaxis.grid(True, color="lightgray", linestyle="-", linewidth=0.8)
    ax.xaxis.grid(False)
    ax.legend()

    label_font = 16
    label_offset = 25

    def annotate_point(x, y, text, color, above: bool):
        position = "bottom" if above else "top"
        offset = label_offset if above else -label_offset
        ax.annotate(
            text,
            (x, y),
            xytext=(0, offset),
            textcoords="offset points",
            color=color,
            ha="center",
            va=position,
            fontsize=label_font,
        )

    for q, va, vb in zip(q_numbers, y_a, y_b):
        has_a = va is not None
        has_b = vb is not None
        if not has_a and not has_b:
            continue
        if has_a and has_b:
            if va >= vb:
                annotate_point(q, va, f"{va:.3f}", color_a, above=True)
                annotate_point(q, vb, f"{vb:.3f}", color_b, above=False)
            else:
                annotate_point(q, va, f"{va:.3f}", color_a, above=False)
                annotate_point(q, vb, f"{vb:.3f}", color_b, above=True)
        elif has_a:
            annotate_point(q, va, f"{va:.3f}", color_a, above=True)
        else:
            annotate_point(q, vb, f"{vb:.3f}", color_b, above=True)

    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.show()
    print(f"Saved {output_path}")


meas_dir = Path(file_path)
meas_file = meas_dir / f"meas#{max_meas_number(meas_dir)}.xlsx"
r_to_fq_file = meas_dir / "R_to_fq.csv"
design_file = meas_dir.parent / "Design_value.xlsx"

if r_to_fq_file.exists():
    predicted_fq = read_csv_q_values(r_to_fq_file)
    measured_fq = read_meas_q_values(meas_file, "Qubit frequency (GHz)", header_row=14)
    q_numbers_fq = sorted(set(predicted_fq) | set(measured_fq))
    plot_frequency_compare(
        q_numbers_fq,
        predicted_fq,
        measured_fq,
        "Predicted",
        "Measured",
        "Predicted vs Measured Qubit Frequency",
        meas_dir / "Fq_predicted_compare.png",
        ylim=(2, 7),
    )
else:
    print("沒有 R_to_fq.csv，跳過 Fq_predicted_compare.png")

if design_file.exists():
    design_fr = read_design_column(design_file, "fr (GHz)")
    measured_fr = read_meas_q_values(meas_file, "Bare resonant frequency (GHz)", header_row=9)
    q_numbers_fr = sorted(set(design_fr) | set(measured_fr))
    plot_frequency_compare(
        q_numbers_fr,
        design_fr,
        measured_fr,
        "Design",
        "Measured",
        "Design vs Measured Resonator Frequency",
        meas_dir / "Fr_design_compare.png",
        ylim=(5, 7),
    )
else:
    print("沒有 Design_value.xlsx，跳過 Fr_design_compare.png")

